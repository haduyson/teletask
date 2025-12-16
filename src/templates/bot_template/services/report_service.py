"""
Report Generation Service
Generate statistical reports in CSV, Excel, and PDF formats
"""

import os
import csv
import logging
import hashlib
import secrets
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from io import BytesIO

import pytz

logger = logging.getLogger(__name__)

TZ = pytz.timezone("Asia/Ho_Chi_Minh")

# Export directory
EXPORT_DIR = Path("/home/botpanel/bots/bot_001/exports")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# Report TTL (72 hours)
REPORT_TTL_HOURS = 72


def generate_report_id() -> str:
    """Generate unique report ID."""
    return secrets.token_hex(16)


def generate_password() -> str:
    """Generate random password for report access."""
    return secrets.token_urlsafe(12)


def hash_password(password: str) -> str:
    """Hash password for storage."""
    return hashlib.sha256(password.encode()).hexdigest()


def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against hash."""
    return hashlib.sha256(password.encode()).hexdigest() == password_hash


async def get_tasks_for_export(
    db,
    user_id: int,
    period_start: Optional[datetime],
    period_end: Optional[datetime],
    task_filter: str = "all",
) -> List[Dict[str, Any]]:
    """
    Get tasks for export based on filter criteria.

    Args:
        db: Database connection
        user_id: User ID
        period_start: Start of period (optional)
        period_end: End of period (optional)
        task_filter: Filter type - 'all', 'created', 'assigned', 'received'

    Returns:
        List of task records
    """
    # Build filter conditions
    filter_conditions = []
    params = [user_id]
    param_idx = 2

    if task_filter == "created":
        filter_conditions.append("t.creator_id = $1")
    elif task_filter == "assigned":
        filter_conditions.append("t.creator_id = $1 AND t.assignee_id != $1")
    elif task_filter == "received":
        filter_conditions.append("t.assignee_id = $1 AND t.creator_id != $1")
    else:  # all
        filter_conditions.append("(t.creator_id = $1 OR t.assignee_id = $1)")

    if period_start:
        filter_conditions.append(f"t.created_at >= ${param_idx}")
        params.append(period_start)
        param_idx += 1

    if period_end:
        filter_conditions.append(f"t.created_at < ${param_idx}")
        params.append(period_end)
        param_idx += 1

    where_clause = " AND ".join(filter_conditions)

    query = f"""
        SELECT
            t.public_id,
            t.content,
            t.status,
            t.priority,
            t.progress,
            t.deadline,
            t.completed_at,
            t.created_at,
            t.updated_at,
            t.is_personal,
            t.group_task_id,
            creator.display_name as creator_name,
            creator.username as creator_username,
            assignee.display_name as assignee_name,
            assignee.username as assignee_username,
            g.title as group_name
        FROM tasks t
        LEFT JOIN users creator ON t.creator_id = creator.id
        LEFT JOIN users assignee ON t.assignee_id = assignee.id
        LEFT JOIN groups g ON t.group_id = g.id
        WHERE t.is_deleted = false AND {where_clause}
        ORDER BY t.created_at DESC
    """

    rows = await db.fetch_all(query, *params)
    return [dict(r) for r in rows]


async def calculate_stats_for_export(
    db,
    user_id: int,
    period_start: Optional[datetime],
    period_end: Optional[datetime],
) -> Dict[str, Any]:
    """Calculate statistics for export report."""
    params = [user_id]
    param_idx = 2

    date_filter = ""
    if period_start:
        date_filter += f" AND created_at >= ${param_idx}"
        params.append(period_start)
        param_idx += 1
    if period_end:
        date_filter += f" AND created_at < ${param_idx}"
        params.append(period_end)
        param_idx += 1

    query = f"""
        SELECT
            COUNT(*) FILTER (WHERE creator_id = $1) as tasks_created,
            COUNT(*) FILTER (WHERE creator_id = $1 AND status = 'completed') as created_completed,
            COUNT(*) FILTER (WHERE creator_id = $1 AND assignee_id != $1) as tasks_assigned,
            COUNT(*) FILTER (WHERE creator_id = $1 AND assignee_id != $1 AND status = 'completed') as assigned_completed,
            COUNT(*) FILTER (WHERE assignee_id = $1 AND creator_id != $1) as tasks_received,
            COUNT(*) FILTER (WHERE assignee_id = $1 AND creator_id != $1 AND status = 'completed') as received_completed,
            COUNT(*) FILTER (WHERE assignee_id = $1 AND is_personal = true) as personal_tasks,
            COUNT(*) FILTER (WHERE assignee_id = $1 AND is_personal = true AND status = 'completed') as personal_completed,
            COUNT(*) FILTER (WHERE (creator_id = $1 OR assignee_id = $1) AND status != 'completed' AND deadline < NOW()) as overdue
        FROM tasks
        WHERE is_deleted = false AND (creator_id = $1 OR assignee_id = $1) {date_filter}
    """

    row = await db.fetch_one(query, *params)

    if row:
        return {
            "tasks_created": row["tasks_created"] or 0,
            "created_completed": row["created_completed"] or 0,
            "tasks_assigned": row["tasks_assigned"] or 0,
            "assigned_completed": row["assigned_completed"] or 0,
            "tasks_received": row["tasks_received"] or 0,
            "received_completed": row["received_completed"] or 0,
            "personal_tasks": row["personal_tasks"] or 0,
            "personal_completed": row["personal_completed"] or 0,
            "overdue": row["overdue"] or 0,
        }
    return {}


def format_status(status: str) -> str:
    """Format status for display."""
    status_map = {
        "pending": "Chờ xử lý",
        "in_progress": "Đang làm",
        "completed": "Hoàn thành",
    }
    return status_map.get(status, status)


def format_priority(priority: str) -> str:
    """Format priority for display."""
    priority_map = {
        "urgent": "Khẩn cấp",
        "high": "Cao",
        "normal": "Bình thường",
        "low": "Thấp",
    }
    return priority_map.get(priority, priority)


def format_datetime_vn(dt: Optional[datetime]) -> str:
    """Format datetime for Vietnamese display."""
    if not dt:
        return "—"
    if dt.tzinfo is None:
        dt = TZ.localize(dt)
    else:
        dt = dt.astimezone(TZ)
    return dt.strftime("%H:%M %d/%m/%Y")


async def generate_csv_report(
    db,
    user_id: int,
    period_start: Optional[datetime],
    period_end: Optional[datetime],
    task_filter: str,
    user_name: str,
) -> Tuple[str, int]:
    """
    Generate CSV report.

    Returns:
        Tuple of (file_path, file_size)
    """
    tasks = await get_tasks_for_export(db, user_id, period_start, period_end, task_filter)
    stats = await calculate_stats_for_export(db, user_id, period_start, period_end)

    report_id = generate_report_id()
    file_path = EXPORT_DIR / f"{report_id}.csv"

    with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow(["BÁO CÁO THỐNG KÊ CÔNG VIỆC"])
        writer.writerow(["Người dùng:", user_name])

        period_str = "Tất cả thời gian"
        if period_start and period_end:
            period_str = f"{format_datetime_vn(period_start)} - {format_datetime_vn(period_end)}"
        writer.writerow(["Khoảng thời gian:", period_str])
        writer.writerow(["Xuất lúc:", format_datetime_vn(datetime.now(TZ))])
        writer.writerow([])

        # Statistics summary
        writer.writerow(["THỐNG KÊ TỔNG HỢP"])
        writer.writerow(["Việc đã tạo:", stats.get("tasks_created", 0)])
        writer.writerow(["  - Hoàn thành:", stats.get("created_completed", 0)])
        writer.writerow(["Việc đã giao:", stats.get("tasks_assigned", 0)])
        writer.writerow(["  - Hoàn thành:", stats.get("assigned_completed", 0)])
        writer.writerow(["Việc được giao:", stats.get("tasks_received", 0)])
        writer.writerow(["  - Hoàn thành:", stats.get("received_completed", 0)])
        writer.writerow(["Việc cá nhân:", stats.get("personal_tasks", 0)])
        writer.writerow(["  - Hoàn thành:", stats.get("personal_completed", 0)])
        writer.writerow(["Việc trễ hạn:", stats.get("overdue", 0)])
        writer.writerow([])

        # Task list
        writer.writerow(["CHI TIẾT CÔNG VIỆC"])
        writer.writerow([
            "Mã việc", "Nội dung", "Trạng thái", "Ưu tiên", "Tiến độ",
            "Người tạo", "Người nhận", "Deadline", "Hoàn thành lúc", "Tạo lúc", "Nhóm"
        ])

        for task in tasks:
            writer.writerow([
                task["public_id"],
                task["content"],
                format_status(task["status"]),
                format_priority(task["priority"]),
                f"{task['progress'] or 0}%",
                task.get("creator_name") or task.get("creator_username") or "—",
                task.get("assignee_name") or task.get("assignee_username") or "—",
                format_datetime_vn(task.get("deadline")),
                format_datetime_vn(task.get("completed_at")),
                format_datetime_vn(task.get("created_at")),
                task.get("group_name") or "—",
            ])

    file_size = file_path.stat().st_size
    return str(file_path), file_size


async def generate_excel_report(
    db,
    user_id: int,
    period_start: Optional[datetime],
    period_end: Optional[datetime],
    task_filter: str,
    user_name: str,
) -> Tuple[str, int]:
    """
    Generate Excel report with charts.

    Returns:
        Tuple of (file_path, file_size)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    tasks = await get_tasks_for_export(db, user_id, period_start, period_end, task_filter)
    stats = await calculate_stats_for_export(db, user_id, period_start, period_end)

    report_id = generate_report_id()
    file_path = EXPORT_DIR / f"{report_id}.xlsx"

    wb = Workbook()

    # ===== Sheet 1: Dashboard =====
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    # Styles
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    number_font = Font(size=14, bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws_dashboard["A1"] = "BÁO CÁO THỐNG KÊ CÔNG VIỆC"
    ws_dashboard["A1"].font = title_font
    ws_dashboard.merge_cells("A1:F1")

    ws_dashboard["A2"] = f"Người dùng: {user_name}"
    period_str = "Tất cả thời gian"
    if period_start and period_end:
        period_str = f"{format_datetime_vn(period_start)} - {format_datetime_vn(period_end)}"
    ws_dashboard["A3"] = f"Khoảng thời gian: {period_str}"
    ws_dashboard["A4"] = f"Xuất lúc: {format_datetime_vn(datetime.now(TZ))}"

    # Statistics summary - starting from row 6
    row = 6
    stats_data = [
        ("Loại", "Tổng", "Hoàn thành", "Tỷ lệ (%)"),
        ("Việc đã tạo", stats.get("tasks_created", 0), stats.get("created_completed", 0),
         round(stats.get("created_completed", 0) / max(stats.get("tasks_created", 1), 1) * 100, 1)),
        ("Việc đã giao", stats.get("tasks_assigned", 0), stats.get("assigned_completed", 0),
         round(stats.get("assigned_completed", 0) / max(stats.get("tasks_assigned", 1), 1) * 100, 1)),
        ("Việc được giao", stats.get("tasks_received", 0), stats.get("received_completed", 0),
         round(stats.get("received_completed", 0) / max(stats.get("tasks_received", 1), 1) * 100, 1)),
        ("Việc cá nhân", stats.get("personal_tasks", 0), stats.get("personal_completed", 0),
         round(stats.get("personal_completed", 0) / max(stats.get("personal_tasks", 1), 1) * 100, 1)),
    ]

    for i, row_data in enumerate(stats_data):
        for j, value in enumerate(row_data):
            cell = ws_dashboard.cell(row=row + i, column=j + 1, value=value)
            cell.border = thin_border
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.alignment = Alignment(horizontal="center") if j > 0 else Alignment(horizontal="left")

    # Adjust column widths
    ws_dashboard.column_dimensions['A'].width = 20
    ws_dashboard.column_dimensions['B'].width = 12
    ws_dashboard.column_dimensions['C'].width = 15
    ws_dashboard.column_dimensions['D'].width = 12

    # Add Pie Chart for task distribution
    chart_row = row + len(stats_data) + 2

    # Pie chart data
    ws_dashboard.cell(row=chart_row, column=6, value="Phân bổ việc")
    ws_dashboard.cell(row=chart_row + 1, column=6, value="Việc đã giao")
    ws_dashboard.cell(row=chart_row + 1, column=7, value=stats.get("tasks_assigned", 0))
    ws_dashboard.cell(row=chart_row + 2, column=6, value="Việc được giao")
    ws_dashboard.cell(row=chart_row + 2, column=7, value=stats.get("tasks_received", 0))
    ws_dashboard.cell(row=chart_row + 3, column=6, value="Việc cá nhân")
    ws_dashboard.cell(row=chart_row + 3, column=7, value=stats.get("personal_tasks", 0))

    if any([stats.get("tasks_assigned", 0), stats.get("tasks_received", 0), stats.get("personal_tasks", 0)]):
        pie_chart = PieChart()
        pie_chart.title = "Phân bổ công việc"
        labels = Reference(ws_dashboard, min_col=6, min_row=chart_row + 1, max_row=chart_row + 3)
        data = Reference(ws_dashboard, min_col=7, min_row=chart_row, max_row=chart_row + 3)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showVal = False
        pie_chart.width = 12
        pie_chart.height = 8
        ws_dashboard.add_chart(pie_chart, "A14")

    # ===== Sheet 2: Task Details =====
    ws_tasks = wb.create_sheet("Chi tiết việc")

    # Headers
    headers = ["Mã việc", "Nội dung", "Trạng thái", "Ưu tiên", "Tiến độ",
               "Người tạo", "Người nhận", "Deadline", "Hoàn thành", "Tạo lúc", "Nhóm"]

    for col, header in enumerate(headers, 1):
        cell = ws_tasks.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Task data
    status_fills = {
        "completed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "in_progress": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "pending": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for row_idx, task in enumerate(tasks, 2):
        row_data = [
            task["public_id"],
            task["content"][:100] + ("..." if len(task["content"]) > 100 else ""),
            format_status(task["status"]),
            format_priority(task["priority"]),
            f"{task['progress'] or 0}%",
            task.get("creator_name") or task.get("creator_username") or "—",
            task.get("assignee_name") or task.get("assignee_username") or "—",
            format_datetime_vn(task.get("deadline")),
            format_datetime_vn(task.get("completed_at")),
            format_datetime_vn(task.get("created_at")),
            task.get("group_name") or "—",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws_tasks.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 2:  # Content column - add text wrap
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 3:  # Status column
                cell.fill = status_fills.get(task["status"], PatternFill())

    # Adjust column widths
    column_widths = [12, 50, 15, 12, 10, 20, 20, 18, 18, 18, 20]
    for i, width in enumerate(column_widths, 1):
        ws_tasks.column_dimensions[chr(64 + i)].width = width

    wb.save(file_path)
    file_size = file_path.stat().st_size
    return str(file_path), file_size


async def generate_pdf_report(
    db,
    user_id: int,
    period_start: Optional[datetime],
    period_end: Optional[datetime],
    task_filter: str,
    user_name: str,
) -> Tuple[str, int]:
    """
    Generate PDF report with charts.

    Returns:
        Tuple of (file_path, file_size)
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend

    tasks = await get_tasks_for_export(db, user_id, period_start, period_end, task_filter)
    stats = await calculate_stats_for_export(db, user_id, period_start, period_end)

    report_id = generate_report_id()
    file_path = EXPORT_DIR / f"{report_id}.pdf"

    # Register Vietnamese-compatible fonts (DejaVu Sans supports UTF-8)
    FONT_PATH = "/usr/share/fonts/truetype/dejavu/"
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', f'{FONT_PATH}DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', f'{FONT_PATH}DejaVuSans-Bold.ttf'))
        FONT_NAME = 'DejaVu'
        FONT_BOLD = 'DejaVu-Bold'
    except Exception as e:
        logger.warning(f"Could not load DejaVu font: {e}, using default")
        FONT_NAME = 'Helvetica'
        FONT_BOLD = 'Helvetica-Bold'

    # Create PDF
    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles with Vietnamese font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=FONT_BOLD,
        fontSize=18,
        spaceAfter=20,
        alignment=1  # Center
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=10,
        spaceAfter=5,
        alignment=1
    )

    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName=FONT_BOLD,
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10
    )

    # Title
    elements.append(Paragraph("BÁO CÁO THỐNG KÊ CÔNG VIỆC", title_style))
    elements.append(Paragraph(f"Người dùng: {user_name}", subtitle_style))

    period_str = "Tất cả thời gian"
    if period_start and period_end:
        period_str = f"{format_datetime_vn(period_start)} - {format_datetime_vn(period_end)}"
    elements.append(Paragraph(f"Khoảng thời gian: {period_str}", subtitle_style))
    elements.append(Paragraph(f"Xuất lúc: {format_datetime_vn(datetime.now(TZ))}", subtitle_style))
    elements.append(Spacer(1, 20))

    # Statistics summary table
    elements.append(Paragraph("THỐNG KÊ TỔNG HỢP", section_style))

    stats_data = [
        ["Loại", "Tổng", "Hoàn thành", "Tỷ lệ (%)"],
        ["Việc đã tạo", str(stats.get("tasks_created", 0)), str(stats.get("created_completed", 0)),
         f"{round(stats.get('created_completed', 0) / max(stats.get('tasks_created', 1), 1) * 100, 1)}%"],
        ["Việc đã giao", str(stats.get("tasks_assigned", 0)), str(stats.get("assigned_completed", 0)),
         f"{round(stats.get('assigned_completed', 0) / max(stats.get('tasks_assigned', 1), 1) * 100, 1)}%"],
        ["Việc được giao", str(stats.get("tasks_received", 0)), str(stats.get("received_completed", 0)),
         f"{round(stats.get('received_completed', 0) / max(stats.get('tasks_received', 1), 1) * 100, 1)}%"],
        ["Việc cá nhân", str(stats.get("personal_tasks", 0)), str(stats.get("personal_completed", 0)),
         f"{round(stats.get('personal_completed', 0) / max(stats.get('personal_tasks', 1), 1) * 100, 1)}%"],
        ["Việc trễ hạn", str(stats.get("overdue", 0)), "—", "—"],
    ]

    stats_table = Table(stats_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D9E2F3')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))

    # Generate charts
    chart_path = EXPORT_DIR / f"{report_id}_chart.png"

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))

    # Pie chart - Task distribution
    labels = ['Việc đã giao', 'Việc được giao', 'Việc cá nhân']
    sizes = [
        stats.get("tasks_assigned", 0),
        stats.get("tasks_received", 0),
        stats.get("personal_tasks", 0)
    ]
    colors_pie = ['#4472C4', '#ED7D31', '#70AD47']

    if sum(sizes) > 0:
        ax1.pie(sizes, labels=labels, colors=colors_pie, autopct='%1.1f%%', startangle=90)
        ax1.set_title('Phân bổ công việc', fontsize=12, fontweight='bold')
    else:
        ax1.text(0.5, 0.5, 'Không có dữ liệu', ha='center', va='center')
        ax1.set_title('Phân bổ công việc', fontsize=12, fontweight='bold')

    # Bar chart - Completion rates
    categories = ['Đã tạo', 'Đã giao', 'Được giao', 'Cá nhân']
    total = [
        stats.get("tasks_created", 0),
        stats.get("tasks_assigned", 0),
        stats.get("tasks_received", 0),
        stats.get("personal_tasks", 0)
    ]
    completed = [
        stats.get("created_completed", 0),
        stats.get("assigned_completed", 0),
        stats.get("received_completed", 0),
        stats.get("personal_completed", 0)
    ]

    x = range(len(categories))
    width = 0.35

    bars1 = ax2.bar([i - width/2 for i in x], total, width, label='Tổng', color='#4472C4')
    bars2 = ax2.bar([i + width/2 for i in x], completed, width, label='Hoàn thành', color='#70AD47')

    ax2.set_ylabel('Số lượng')
    ax2.set_title('Tỷ lệ hoàn thành', fontsize=12, fontweight='bold')
    ax2.set_xticks(x)
    ax2.set_xticklabels(categories, fontsize=9)
    ax2.legend()

    plt.tight_layout()
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()

    # Add chart image
    elements.append(Image(str(chart_path), width=16*cm, height=6.5*cm))
    elements.append(Spacer(1, 20))

    # Task list (first 30 tasks)
    if tasks:
        elements.append(Paragraph("CHI TIẾT CÔNG VIỆC", section_style))

        # Content style for wrapping text
        content_style = ParagraphStyle(
            'ContentStyle',
            parent=styles['Normal'],
            fontName=FONT_NAME,
            fontSize=8,
            leading=10,
        )

        task_data = [["Mã", "Nội dung", "Trạng thái", "Deadline"]]
        for task in tasks[:30]:
            content = task["content"][:60] + ("..." if len(task["content"]) > 60 else "")
            task_data.append([
                task["public_id"],
                Paragraph(content, content_style),
                format_status(task["status"]),
                format_datetime_vn(task.get("deadline"))[:10] if task.get("deadline") else "—"
            ])

        if len(tasks) > 30:
            task_data.append(["...", Paragraph(f"và {len(tasks) - 30} việc khác", content_style), "", ""])

        task_table = Table(task_data, colWidths=[2*cm, 9*cm, 3*cm, 3*cm])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
            ('FONTNAME', (0, 1), (0, -1), FONT_NAME),
            ('FONTNAME', (2, 1), (3, -1), FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (0, -1), 8),
            ('FONTSIZE', (2, 1), (3, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))
        elements.append(task_table)

    # Build PDF
    doc.build(elements)

    # Clean up chart image
    if chart_path.exists():
        chart_path.unlink()

    file_size = file_path.stat().st_size
    return str(file_path), file_size


async def create_export_report(
    db,
    user_id: int,
    user_name: str,
    report_type: str,
    file_format: str,
    task_filter: str,
    period_start: Optional[datetime] = None,
    period_end: Optional[datetime] = None,
) -> Dict[str, Any]:
    """
    Create and store an export report.

    Args:
        db: Database connection
        user_id: User ID
        user_name: User display name
        report_type: 'last7', 'last30', 'this_week', 'last_week', 'this_month', 'last_month', 'custom'
        file_format: 'csv', 'xlsx', 'pdf'
        task_filter: 'all', 'created', 'assigned', 'received'
        period_start: Start of custom period
        period_end: End of custom period

    Returns:
        Dict with report info including URL and password
    """
    now = datetime.now(TZ)

    # Calculate period based on report_type
    if report_type == "last7":
        period_start = now - timedelta(days=7)
        period_end = now
    elif report_type == "last30":
        period_start = now - timedelta(days=30)
        period_end = now
    elif report_type == "this_week":
        period_start = now - timedelta(days=now.weekday())
        period_start = period_start.replace(hour=0, minute=0, second=0, microsecond=0)
        period_end = now
    elif report_type == "last_week":
        period_end = now - timedelta(days=now.weekday())
        period_end = period_end.replace(hour=0, minute=0, second=0, microsecond=0)
        period_start = period_end - timedelta(days=7)
    elif report_type == "this_month":
        period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        period_end = now
    elif report_type == "last_month":
        period_end = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if period_end.month == 1:
            period_start = period_end.replace(year=period_end.year - 1, month=12)
        else:
            period_start = period_end.replace(month=period_end.month - 1)
    # else: use provided period_start and period_end

    # Generate report based on format
    if file_format == "csv":
        file_path, file_size = await generate_csv_report(
            db, user_id, period_start, period_end, task_filter, user_name
        )
    elif file_format == "xlsx":
        file_path, file_size = await generate_excel_report(
            db, user_id, period_start, period_end, task_filter, user_name
        )
    elif file_format == "pdf":
        file_path, file_size = await generate_pdf_report(
            db, user_id, period_start, period_end, task_filter, user_name
        )
    else:
        raise ValueError(f"Unknown file format: {file_format}")

    # Generate report ID and password
    report_id = generate_report_id()
    password = generate_password()
    password_hash = hash_password(password)
    expires_at = now + timedelta(hours=REPORT_TTL_HOURS)

    # Store in database
    await db.execute(
        """
        INSERT INTO export_reports (
            report_id, password_hash, user_id, report_type, file_format,
            task_filter, period_start, period_end, file_path, file_size, expires_at
        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11)
        """,
        report_id, password_hash, user_id, report_type, file_format,
        task_filter, period_start, period_end, file_path, file_size, expires_at
    )

    return {
        "report_id": report_id,
        "password": password,
        "file_format": file_format,
        "file_size": file_size,
        "expires_at": expires_at,
        "period_start": period_start,
        "period_end": period_end,
    }


async def get_report_by_id(db, report_id: str) -> Optional[Dict[str, Any]]:
    """Get report record by ID."""
    row = await db.fetch_one(
        "SELECT * FROM export_reports WHERE report_id = $1",
        report_id
    )
    return dict(row) if row else None


async def increment_download_count(db, report_id: str) -> None:
    """Increment download count for a report."""
    await db.execute(
        "UPDATE export_reports SET download_count = download_count + 1 WHERE report_id = $1",
        report_id
    )


async def cleanup_expired_reports(db) -> int:
    """Delete expired reports from database and filesystem."""
    now = datetime.now(TZ)

    # Get expired reports
    rows = await db.fetch_all(
        "SELECT id, file_path FROM export_reports WHERE expires_at < $1",
        now
    )

    deleted_count = 0
    for row in rows:
        try:
            file_path = Path(row["file_path"])
            if file_path.exists():
                file_path.unlink()
            deleted_count += 1
        except Exception as e:
            logger.warning(f"Failed to delete report file: {e}")

    # Delete from database
    await db.execute("DELETE FROM export_reports WHERE expires_at < $1", now)

    return deleted_count
